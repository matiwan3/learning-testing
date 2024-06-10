from faker import Faker
from openpyxl import Workbook

wb = Workbook()  # Workbook is the container for all other parts of the document.
ws = wb.active  # Get the currently active sheet or None
fake_data = Faker('pl_PL')  # set the language to pl

for i in range(1, 11):
    for j in range(1,4):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email() 
        ws.cell(row=i, column=3).value = fake_data.address()
wb.save("testData.csv") 